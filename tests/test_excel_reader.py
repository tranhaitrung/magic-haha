"""Tests for excel_reader module."""

import unittest
from pathlib import Path
from tempfile import NamedTemporaryFile

from src.excel_reader import InputValidationError, read_channel_tasks


class TestExcelReader(unittest.TestCase):
    def test_read_valid_excel(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")

        # Create a temp Excel file
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "channel_url"
        ws["B1"] = "note"
        ws["A2"] = "https://www.facebook.com/testpage"
        ws["B2"] = "Test page"
        ws["A3"] = "https://www.facebook.com/groups/123"
        ws["B3"] = "Test group"

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        try:
            tasks = read_channel_tasks(tmp_path)
            self.assertEqual(len(tasks), 2)
            self.assertEqual(tasks[0].channel_url, "https://www.facebook.com/testpage")
            self.assertEqual(tasks[0].note, "Test page")
            self.assertEqual(tasks[1].channel_url, "https://www.facebook.com/groups/123")
        finally:
            tmp_path.unlink()

    def test_read_no_channel_url_header(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "url"
        ws["A2"] = "https://www.facebook.com/test"

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        try:
            with self.assertRaises(InputValidationError):
                read_channel_tasks(tmp_path)
        finally:
            tmp_path.unlink()

    def test_read_dedupe_urls(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "channel_url"
        ws["A2"] = "https://www.facebook.com/test"
        ws["A3"] = "https://www.facebook.com/test"  # Duplicate
        ws["A4"] = "https://www.facebook.com/test/"  # Trailing slash

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        try:
            tasks = read_channel_tasks(tmp_path)
            self.assertEqual(len(tasks), 1)
        finally:
            tmp_path.unlink()

    def test_read_invalid_urls(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "channel_url"
        ws["A2"] = "not a url"
        ws["A3"] = ""
        ws["A4"] = "https://valid.com/page"

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        try:
            tasks = read_channel_tasks(tmp_path)
            self.assertEqual(len(tasks), 1)
            self.assertEqual(tasks[0].channel_url, "https://valid.com/page")
        finally:
            tmp_path.unlink()

    def test_read_file_not_found(self) -> None:
        with self.assertRaises(InputValidationError):
            read_channel_tasks(Path("/nonexistent/file.xlsx"))

    def test_read_wrong_format(self) -> None:
        with NamedTemporaryFile(suffix=".txt", delete=False) as tmp:
            tmp_path = Path(tmp.name)
            tmp_path.write_text("channel_url\nhttps://test.com")

        try:
            with self.assertRaises(InputValidationError):
                read_channel_tasks(tmp_path)
        finally:
            tmp_path.unlink()


if __name__ == "__main__":
    unittest.main()
