from __future__ import annotations

from pathlib import Path
from urllib.parse import urlparse

from src.models import ChannelTask


class InputValidationError(ValueError):
    pass


def _is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def read_channel_tasks(input_file: Path) -> list[ChannelTask]:
    if not input_file.exists():
        raise InputValidationError(f"Input file not found: {input_file}")
    if input_file.suffix.lower() != ".xlsx":
        raise InputValidationError("Input file must be .xlsx")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl") from exc

    wb = load_workbook(input_file, read_only=True, data_only=True)
    sheet = wb.active

    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        headers.append((str(cell.value).strip().lower() if cell.value is not None else ""))

    if "channel_url" not in headers:
        raise InputValidationError("Missing required header: channel_url")

    url_col_idx = headers.index("channel_url")
    note_col_idx = headers.index("note") if "note" in headers else None

    tasks: list[ChannelTask] = []
    seen: set[str] = set()

    for row in sheet.iter_rows(min_row=2):
        raw_url = row[url_col_idx].value
        if raw_url is None:
            continue

        channel_url = str(raw_url).strip()
        if not channel_url or not _is_http_url(channel_url):
            continue

        normalized = channel_url.rstrip("/")
        if normalized in seen:
            continue

        seen.add(normalized)
        note = None
        if note_col_idx is not None and row[note_col_idx].value is not None:
            note = str(row[note_col_idx].value).strip()

        tasks.append(ChannelTask(channel_url=normalized, note=note))

    return tasks
